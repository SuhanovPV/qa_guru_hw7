import csv
import os
from zipfile import ZipFile, ZIP_DEFLATED
from pytest import fixture
from pypdf import PdfReader
from openpyxl import load_workbook

RESOURCES = os.path.join(os.path.dirname(os.path.abspath(__name__)), 'files')
PDF_FILE = 'pdf_file.pdf'
CSV_FILE = 'csv_file.csv'
XLSX_FILE = 'xlsx_file.xlsx'
ARCHIVE = 'archive.zip'
FILES = [PDF_FILE, CSV_FILE, XLSX_FILE]


@fixture()
def create_archive():
    with ZipFile(os.path.join(RESOURCES, ARCHIVE), mode='w', compression=ZIP_DEFLATED) as my_zip:
        for file in FILES:
            my_zip.write(os.path.join(RESOURCES, file), file)


def test_zip_archive():
    with ZipFile(os.path.join(RESOURCES, ARCHIVE), mode='w', compression=ZIP_DEFLATED) as my_zip:
        for file in FILES:
            my_zip.write(os.path.join(RESOURCES, file), file)

    with ZipFile(os.path.join(RESOURCES, ARCHIVE), mode='r') as my_zip:
        for file in my_zip.namelist():
            assert file in FILES


def test_pdf(create_archive):
    expected_text = PdfReader(os.path.join(RESOURCES, PDF_FILE)).pages[0].extract_text()
    with ZipFile(os.path.join(RESOURCES, ARCHIVE), mode='r') as my_zip:
        with my_zip.open(PDF_FILE, mode='r') as archive_pdf:
            archive_pdf_text = PdfReader(archive_pdf).pages[0].extract_text()
            assert archive_pdf_text == expected_text


def test_csv(create_archive):
    with open(os.path.join(RESOURCES, CSV_FILE), 'rb') as csv_file:
        expected_content = csv_file.read().decode('utf-8')
        expected_data = csv.reader(expected_content.splitlines())

    with ZipFile(os.path.join(RESOURCES, ARCHIVE), mode='r') as my_zip:
        with my_zip.open(CSV_FILE) as archive_csv:
            content = archive_csv.read().decode('utf-8')
            archive_data = csv.reader(content.splitlines())

    for archive_value, expected_value in zip(archive_data, expected_data):
        assert archive_value == expected_value


def test_xlsx(create_archive):
    with ZipFile(os.path.join(RESOURCES, ARCHIVE), mode='r') as my_zip:
        with my_zip.open(XLSX_FILE, mode='r') as archive_xlsx:
            archive_ws = load_workbook(archive_xlsx).active
            archive_data = archive_ws['A']

            expected_ws = load_workbook(os.path.join(RESOURCES, XLSX_FILE)).active
            expected_data = expected_ws['A']

            for archive_cell, expected_cell in zip(archive_data, expected_data):
                assert archive_cell.value == expected_cell.value
